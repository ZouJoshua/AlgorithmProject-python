#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Author  : Joshua
@Time    : 2019/3/22 11:23
@File    : regional_process.py
@Desc    : 地域性服务
"""

import os
from os.path import dirname
import sys
root_path = dirname(dirname(dirname(dirname(os.path.realpath(__file__)))))
root_nlp_path = dirname(dirname(dirname(os.path.realpath(__file__))))
sys.path.append(root_path)
sys.path.append(root_nlp_path)

import json
import re
import xlrd
import random
from openpyxl import load_workbook


def regional_preprocess_from_xlsx(sheetname='town'):
    town_xlsx_file = os.path.join(root_nlp_path, 'data', 'india_town.xlsx')
    wb = load_workbook(town_xlsx_file)
    # sheets = wb.sheetnames
    # print(sheets)
    sheet = wb[sheetname]
    rows = sheet.rows
    cols = sheet.columns
    result = dict()

    _c = 0
    for row in rows:
        _c += 1
        if _c > 1:
            line = [col.value for col in row]
            if len(line) == 8:
                states = line[1].title().replace('&', 'and')
                district = line[3].title()
                sub_district = re.sub(' \d+', '', line[5].title())
                town = re.sub(' \(.*?\)', '', line[7].title())
                if states in result.keys():
                    if district not in result[states]['districts']:
                        result[states]['districts'].append(district)
                    if sub_district not in result[states]['sub_districts']:
                        result[states]['sub_districts'].append(sub_district)
                    if town not in result[states]['town']:
                        result[states]['town'].append(town)
                else:
                    result[states] = dict()
                    result[states]['districts'] = list()
                    result[states]['sub_districts'] = list()
                    result[states]['town'] = list()
                    result[states]['districts'].append(district)
                    result[states]['sub_districts'].append(sub_district)
                    result[states]['town'].append(town)
    outfile = os.path.join(root_nlp_path, 'data', 'india_division_new.json')
    with open(outfile, 'w', encoding='utf-8') as wf:
        for k, v in result.items():
            out = dict()
            out[k] = v
            wf.write(json.dumps(out) + '\n')
    return

## 修复地名问题
"""
1.去除掉部分与常用词相同的单词，如：Bank、Kim
"""
def regional_preprocess(filename='india_division.json'):
    regi_file = os.path.join(root_nlp_path, 'data', filename)
    reader = open(regi_file, 'r', encoding='utf-8')
    line = json.load(reader)
    out = dict()
    # print(line)
    for k, v in line.items():
        if k in ['States', 'Union territories']:
            states = line[k]
            for _k, _v in states.items():
                out_result = {"regional": '',
                              'is_capital': False, 'is_regions': False,
                              'is_divisions': False, 'is_districts': False,
                              'is_headquarters': False, 'is_largest_city': False,
                              'is_sub_districts': False, 'is_town': False}
                if _k not in out.keys():
                    out_result['regional'] = _k
                    out[_k] = out_result
                for i, j in _v.items():
                    if j:
                        if i == 'districts_new':
                            i = 'districts'
                            j_upper = list()
                            for d in j:
                                j_upper.append(d.title().replace("District", "").strip().upper())
                            j += j_upper
                        elif i == 'capital' or i == 'headquarters':
                            j_upper = list()
                            for d in j:
                                j_upper.append(d.title().strip().upper())
                            j += j_upper
                        for _j in j:
                            if _j not in out.keys():
                                out_result = {"regional": '',
                                              'is_capital': False, 'is_regions': False,
                                              'is_divisions': False, 'is_districts': False,
                                              'is_headquarters': False, 'is_largest_city': False,
                                              'is_sub_districts': False, 'is_town': False}
                                out_result['regional'] = _k
                                out_result['is_{}'.format(i)] = True
                                _j_clean = re.sub('\s+', ' ', _j.replace("*", " ").replace("District", "").replace("Division", "").strip())
                                out[_j_clean] = out_result
                            else:
                                out_result = out[_j]
                                out_result['regional'] = _k
                                out_result['is_{}'.format(i)] = True
                                _j_clean = re.sub('\s+', ' ', _j.replace("*", " ").replace("District", "").replace("Division", "").strip())
                                out[_j_clean] = out_result
                    else:
                        continue
    reader.close()
    regi2map_file = os.path.join(root_nlp_path, 'data', 'india_names2regions.json')
    with open(regi2map_file, 'w') as f:
        json.dump(out, f, indent=4)
    return


def read_map_file(mapfile='india_names2regions.json'):
    map_file = os.path.join(root_nlp_path, 'data', mapfile)
    if not os.path.exists(map_file):
        regional_preprocess()
    else:
        pass
    with open(map_file, 'r', encoding='utf-8') as reader:
        names_map = json.load(reader)
    return names_map

def get_regional(text):
    names_map = read_map_file()
    out_count = get_detail_regional(text, names_map)
    result = _count_regional(out_count, names_map)
    return result

def get_detail_regional(text, names_map):
    out = dict()
    for key, value in names_map.items():
        # if len(key) < 5:
        #     key += ' '
        all = re.findall(key + '[\W]', text)
        if len(all):
            out[key.strip()] = len(all)
    return _re_count_regional(out)

def _re_count_regional(regional_ct):
    out = dict()
    if regional_ct:
        for k, v in regional_ct.items():
            if k.isupper():
                k = k.title()
            if k not in out:
                out[k] = v
            else:
                out[k] += v
    return out


def _count_regional(result, names_map):
    out = dict()
    for k, v in result.items():
        if k in ['Mumbai', 'Bengaluru', 'Kolkata', 'Hyderabad']:
            region = k
        else:
            region = names_map[k]['regional']
        if region not in out:
            out[region] = v
        else:
            out[region] += v
    return out


def predict_regional(regional_ct, text, topk=3):
    regional = dict()
    regional['regional'] = list()
    if regional_ct:
        ks = regional_ct.keys()
        deh = ['National Capital Territory Of Delhi', 'Delhi']
        dehil_names = set(deh) & set(ks)
        city = ['Mumbai', 'Bengaluru', 'Kolkata', 'Hyderabad']
        city_names = set(city) & set(ks)
        if dehil_names:
            if re.findall('New Delhi:', text) or re.findall('Delhi:', text) or re.findall('DELHI:', text):
                regional['regional'].append('Delhi')
            else:
                if len(dehil_names) == 2:
                    if regional_ct['Delhi'] >= regional_ct['National Capital Territory Of Delhi']:
                        del regional_ct['National Capital Territory Of Delhi']
                    else:
                        del regional_ct['Delhi']
        elif city_names:
            for city in city_names:
                _city = city + ':'
                if re.findall(_city, text) or re.findall(_city.upper(), text):
                    regional['regional'].append(city)
        if not regional['regional']:
            sort_regional_ct = sorted(regional_ct.items(), key=lambda x: x[1], reverse=True)
            topk_regional_ct = sort_regional_ct[:topk]
            if len(topk_regional_ct) == 1:
                regional['regional'].append(topk_regional_ct[0][0])
            elif len(topk_regional_ct) > 1:
                if topk_regional_ct[0][1] > topk_regional_ct[1][1]:
                    regional['regional'].append(topk_regional_ct[0][0])
                elif topk_regional_ct[0][1] == topk_regional_ct[1][1]:
                    regional['regional'].append(topk_regional_ct[0][0])
                    regional['regional'].append(topk_regional_ct[1][0])
                    if len(topk_regional_ct) > 2 and topk_regional_ct[1][1] == topk_regional_ct[2][1]:
                        regional['regional'].append(topk_regional_ct[2][0])
                    else:
                        pass
    return regional


def get_article(file, limit=100):
    reader = open(file, 'r', encoding='utf-8')
    i = 0
    while True:
        line = json.loads(reader.readline())
        yield line
        i += 1
        if i > limit:
            reader.close()
            break

def get_article_random_check(file, outfile, limit=100):
    names_map = read_map_file()
    reader = open(file, 'r', encoding='utf-8')
    lines = reader.readlines()
    # random.shuffle(lines)
    with open(outfile, 'w', encoding='utf-8') as wf:
        for i in lines[:limit]:
            data = json.loads(i)
            text = data['title'] + '.' + data['content']
            out_count = get_detail_regional(text, names_map)
            data['regional_keywords'] = out_count
            data['regional'] = _count_regional(out_count, names_map)
            print(out_count)
            # print(data['regional'])
            print("结果{}".format(predict_regional(data['regional'], text)))
            # wf.write(json.dumps(data)+'\n')
    reader.close()
    return

if __name__ == '__main__':
    # regional_preprocess()
    # regional_preprocess_from_xlsx()
    # datafile = os.path.join(root_path, 'data', 'national')
    # # line = get_article(datafile)
    # # regional_names_map = read_map_file()
    # # s = 0
    # # for i in line:
    # #     s += 1
    # #     if s == 2:
    # #         out = get_detail_regional(i['content'], regional_names_map)
    # #         print(out)
    datafile = os.path.join(root_path, 'data', 'test')
    outfile = os.path.join(root_path, 'data', 'test1')
    get_article_random_check(datafile, outfile)